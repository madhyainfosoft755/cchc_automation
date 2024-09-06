# Utilities/utilities.py

from openpyxl import load_workbook
import csv
import logging
import os
import time

# Configure logging
logging.basicConfig(
    filename='logs/automation.log',
    level=logging.INFO,
    format='%(asctime)s:%(levelname)s:%(message)s'
)
logger = logging.getLogger()

class Utilities:

    @staticmethod
    def read_data_from_excel(filename, sheet):
        datalist = []
        try:
            wb = load_workbook(filename=filename)
            sh = wb[sheet]
            row_count = sh.max_row
            col_count = sh.max_column

            for i in range(2, row_count + 1):
                row = []
                for j in range(1, col_count + 1):
                    row.append(sh.cell(row=i, column=j).value)
                datalist.append(row)
            logger.info(f"Data read from Excel: {filename}, Sheet: {sheet}")
            return datalist
        except Exception as e:
            logger.error(f"Failed to read Excel data: {e}")
            return []

    @staticmethod
    def read_data_from_CSV(filename):
        datalist = []
        try:
            with open(filename, "r", newline='') as csvdata:
                reader = csv.reader(csvdata)
                next(reader)  # Skip header
                for rows in reader:
                    datalist.append(rows)
            logger.info(f"Data read from CSV: {filename}")
            return datalist
        except Exception as e:
            logger.error(f"Failed to read CSV data: {e}")
            return []

    @staticmethod
    def take_screenshot(driver, name="screenshot"):
        timestamp = time.strftime("%Y%m%d-%H%M%S")
        screenshot_name = f"{name}_{timestamp}.png"
        screenshot_path = os.path.join('screenshots', screenshot_name)
        try:
            driver.save_screenshot(screenshot_path)
            logger.info(f"Screenshot taken: {screenshot_path}")
        except Exception as e:
            logger.error(f"Failed to take screenshot: {e}")
